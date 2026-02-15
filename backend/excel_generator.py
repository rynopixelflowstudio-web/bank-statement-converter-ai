from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
import os


class ExcelGenerator:
    """Generate formatted Excel spreadsheet from transaction data"""
    
    def __init__(self, transactions: List[Dict], output_dir: str = "outputs"):
        self.transactions = transactions
        self.output_dir = output_dir
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
    
    def generate(self, filename: str = "transactions.xlsx") -> str:
        """Generate Excel file and return the file path"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Transactions"
        
        # Define column headers
        headers = ["Date", "Description", "Reference", "Debit", "Credit", "Balance"]
        ws.append(headers)
        
        # Style the header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add transaction data
        for transaction in self.transactions:
            row = [
                transaction.get('date', ''),
                transaction.get('description', ''),
                transaction.get('reference', ''),
                transaction.get('debit', ''),
                transaction.get('credit', ''),
                transaction.get('balance', '')
            ]
            ws.append(row)
        
        # Apply formatting to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Don't center-align header
                    cell.alignment = Alignment(vertical="center")
        
        # Auto-size columns
        column_widths = {
            'A': 12,  # Date
            'B': 50,  # Description
            'C': 15,  # Reference
            'D': 12,  # Debit
            'E': 12,  # Credit
            'F': 12   # Balance
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Format amount columns as currency (right-aligned)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Add number format if cell has a value
                if cell.value:
                    try:
                        # Try to convert to float for proper number formatting
                        float_val = float(str(cell.value).replace(',', ''))
                        cell.value = float_val
                        cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        pass
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Save the workbook
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def generate_with_timestamp(self) -> str:
        """Generate Excel file with timestamp in filename"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bank_transactions_{timestamp}.xlsx"
        return self.generate(filename)
